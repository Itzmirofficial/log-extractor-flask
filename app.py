import os
import csv
import pandas as pd
import re
from datetime import datetime, timedelta
from flask import Flask, render_template, request, send_file, redirect, url_for, flash, jsonify
from werkzeug.utils import secure_filename
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import tempfile
import shutil
import codecs
import logging
import time

# Set up logging
logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

app = Flask(__name__)
app.secret_key = "employee_log_extractor_secret_key"

UPLOAD_FOLDER = os.path.join(tempfile.gettempdir(), 'employee_log_extractor')
OUTPUT_FOLDER = os.path.join(UPLOAD_FOLDER, 'output')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['OUTPUT_FOLDER'] = OUTPUT_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024
app.config['ALLOWED_EXTENSIONS'] = {'csv'}

# List of departments
DEPARTMENTS = [
    "PATHOLOGY", "PHARMACOLOGY", "PHYSIOLOGY", "PSYCHIATRY", "RADIATION ONCOLOGY", "RADIO-DIAGNOSIS",
    "RESPIRATORY MEDICINE", "UG-Students", "M.Ch - Surgical Gastroenterology/G.I. Surgery",
    "M.Ch - Urology/Genito-Urinary Surgery", "MICROBIOLOGY", "OBSTETRICS & GYNAECOLOGY",
    "OPHTHALMOLOGY", "ORTHOPAEDICS", "PAEDIATRICS", "GENERAL MEDICINE", "GENERAL SURGERY", "IT",
    "M.Ch - Cardio Thoracic and Vascular Surgery", "M.Ch - Neuro Surgery", "M.Ch - Paediatric Surgery",
    "M.Ch - Plastic Surgery/Plastic & Reconstructive Surgery", "DM - Cardiology", "DM - Medical Gastroenterology",
    "DM - Nephrology", "DM - Neurology", "EMERGENCY MEDICINE", "ENT/ Otorhinolaryngology",
    "FORENSIC MEDICINE & TOXICOLOGY", "ADMINISTRATION", "ANATOMY", "ANESTHESIOLOGY", "BIOCHEMISTRY",
    "COMMUNITY MEDICINE", "DENTISTRY", "DERMATOLOGY"
]

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']

def format_datetime(value):
    if not value or not isinstance(value, str):
        return "", ""
    try:
        dt = datetime.strptime(value.strip(), "%Y-%m-%d %H:%M:%S")
        return dt.date().strftime("%B %d, %Y"), dt.time().strftime("%H:%M:%S")
    except Exception as e:
        logger.warning(f"Failed to parse datetime '{value}': {e}")
        return "", ""

def time_to_seconds(time_str):
    if not time_str or not re.match(r'^\d{2}:\d{2}:\d{2}$', time_str):
        return 0
    h, m, s = map(int, time_str.split(':'))
    return h * 3600 + m * 60 + s

def time_to_excel_format(time_str):
    if not time_str or time_str.strip() == "":
        return ""
    try:
        if re.match(r'^\d{2}:\d{2}:\d{2}$', time_str):
            return time_str
        dt = datetime.strptime(time_str.strip(), "%H:%M:%S")
        return dt.strftime("%H:%M:%S")
    except Exception as e:
        logger.warning(f"Failed to format time '{time_str}': {e}")
        return ""

def read_csv_safely(file_path):
    encodings = ['utf-8', 'utf-8-sig', 'latin1', 'iso-8859-1', 'cp1252']
    start_time = time.time()
    timeout = 30
    
    for encoding in encodings:
        try:
            with open(file_path, 'r', encoding=encoding) as f:
                reader = csv.reader(f)
                data = list(reader)
                logger.debug(f"Successfully read {file_path} with {encoding}")
                return data, encoding
        except UnicodeDecodeError:
            if time.time() - start_time > timeout:
                logger.error(f"Timeout while reading {file_path}")
                raise TimeoutError("CSV reading timed out")
            continue
    logger.error(f"Unable to decode {file_path} with any encoding")
    raise UnicodeDecodeError("Unable to decode the file with any of the attempted encodings", "", 0, 0, "")

def extract_employees_from_csv(file_paths):
    employees = []
    for file_path in file_paths:
        logger.info(f"Processing employee extraction for {file_path}")
        try:
            reader, _ = read_csv_safely(file_path)
            for row in reader:
                row_text = ' '.join(row)
                if "att-id:" in row_text.lower():
                    header_text = row_text
                    id_match = re.search(r'Att-ID:(\d+)', header_text, re.IGNORECASE)
                    employee_id = id_match.group(1) if id_match else ""
                    
                    name_match = re.search(r'^([^A-Z]*?)(?:\s+Att-ID|\s+Emp)', header_text, re.IGNORECASE)
                    employee_name = name_match.group(1).strip() if name_match else ""
                    if not employee_name:
                        name_match = re.search(r'^(.*?)\s*Att-ID', header_text, re.IGNORECASE)
                        employee_name = name_match.group(1).strip() if name_match else ""
                    if not employee_name:
                        employee_name = header_text.split("Att-ID")[0].strip() if "Att-ID" in header_text else ""
                    
                    if employee_name or employee_id:
                        is_duplicate = any((employee_name and emp['name'] == employee_name) or \
                                          (employee_id and emp['id'] == employee_id) for emp in employees)
                        if not is_duplicate:
                            employees.append({
                                'name': employee_name,
                                'id': employee_id,
                                'display': f"{employee_name} (ID: {employee_id})" if employee_name and employee_id else \
                                          employee_name if employee_name else f"ID: {employee_id}"
                            })
                            logger.debug(f"Added employee: {employee_name} (ID: {employee_id})")
        except Exception as e:
            logger.error(f"Error processing {file_path} for employees: {e}")
    
    employees.sort(key=lambda x: x['name'] if x['name'] else x['id'])
    logger.info(f"Extracted {len(employees)} unique employees")
    return employees


def extract_employee_logs(file_paths, identifiers, search_by, output_format='xlsx', department=None):
    """
    Extract employee attendance logs and generate reports in various formats.
    Fixed version that ensures complete date ranges and proper status handling.
    """
    from datetime import datetime, timedelta
    import pandas as pd
    import os
    import re
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    import logging

    logger = logging.getLogger(__name__)
    
    log_results = []
    wb = Workbook()
    wb.remove(wb.active)
    any_data_found = False
    output_files = {'xlsx': None, 'csv': [], 'html': []}
    display_names = {}
    min_date = None
    max_date = None
    
    # Find the month in the uploaded CSV files first
    csv_month = None
    csv_year = None
    
    # First scan to determine the month from the data
    for file_path in file_paths:
        try:
            reader, encoding = read_csv_safely(file_path)
            for row in reader:
                row_text = ' '.join(row).lower()
                if 'date' in row_text:
                    continue
                # Look for date patterns like "March 15, 2025"
                date_match = re.search(r'([a-z]+)\s+\d+,\s+(\d{4})', row_text, re.IGNORECASE)
                if date_match:
                    month_name = date_match.group(1).capitalize()
                    year = int(date_match.group(2))
                    month_num = {'January': 1, 'February': 2, 'March': 3, 'April': 4, 
                                'May': 5, 'June': 6, 'July': 7, 'August': 8, 
                                'September': 9, 'October': 10, 'November': 11, 'December': 12}.get(month_name)
                    if month_num:
                        csv_month = month_num
                        csv_year = year
                        log_results.append(f"[📅] Detected month from CSV: {month_name} {year}")
                        break
            if csv_month:
                break
        except Exception as e:
            logger.error(f"Error scanning {file_path} for month: {e}")
    
    # If we couldn't determine month from data, use current month as fallback
    if not csv_month:
        today = datetime.now()
        csv_month = today.month
        csv_year = today.year
        log_results.append(f"[📅] No month detected in CSV, using current month: {csv_month}/{csv_year}")
    
    # Set the date range for the report based on the detected month
    report_month_start = datetime(csv_year, csv_month, 1)
    if csv_month == 12:
        report_month_end = datetime(csv_year + 1, 1, 1) - timedelta(days=1)
    else:
        report_month_end = datetime(csv_year, csv_month + 1, 1) - timedelta(days=1)
    
    for identifier in identifiers:
        logger.info(f"Processing logs for identifier: {identifier} ({search_by})")
        all_data_frames = []
        employee_name = ""
        employee_id = ""
        designation = "Senior Resident Ng"  # Default designation
        sheet_name = f"Report_{identifier[:31]}"  # Default sheet name, truncated to 31 chars
        
        for file_path in file_paths:
            logger.debug(f"Scanning {file_path} for {identifier}")
            try:
                reader, encoding = read_csv_safely(file_path)
                log_results.append(f"🔍 Scanning {os.path.basename(file_path)} for {identifier}...")
                log_results.append(f"  Using encoding: {encoding}")
                
                emp_index = None
                search_pattern = identifier.lower() if search_by == 'name' else f"att-id:{identifier.lower()}"
                
                for idx, row in enumerate(reader):
                    row_text = ' '.join(row).lower()
                    if search_pattern in row_text:
                        emp_index = idx
                        header_text = ' '.join(row)
                        if search_by == 'name':
                            employee_name = identifier
                            id_match = re.search(r'Att-ID:(\d+)', header_text, re.IGNORECASE)
                            employee_id = id_match.group(1) if id_match else ""
                        else:
                            employee_id = identifier
                            name_match = re.search(r'^([^A-Z]*?)\s+Att-ID', header_text, re.IGNORECASE)
                            employee_name = name_match.group(1).strip() if name_match else ""
                        logger.debug(f"Found employee: {employee_name} (ID: {employee_id})")
                        break
                
                if emp_index is not None:
                    raw_header = reader[emp_index + 1] if emp_index + 1 < len(reader) else []
                    clean_header = [h.strip().lower().replace(" ", "_") for h in raw_header]
                    data_rows = []
                    
                    for row in reader[emp_index + 2:]:
                        if not row or not any(cell.strip() for cell in row):
                            break
                        if row and len(row) > 0 and search_pattern not in ' '.join(row).lower() and "att-id" in ' '.join(row).lower():
                            break
                        if row and len(row) == len(raw_header):
                            data_rows.append(row)
                    
                    if data_rows:
                        df = pd.DataFrame(data_rows, columns=clean_header)
                        
                        if 'in_time' in df.columns and 'out_time' in df.columns:
                            df['date'], df['in_time_val'] = zip(*df['in_time'].map(format_datetime))
                            df['out_time_val'] = df['out_time'].map(lambda x: format_datetime(x)[1])
                            df.drop(['in_time', 'out_time'], axis=1, inplace=True)
                            
                            time_columns = ['in_time_val', 'out_time_val', 'in_time_short_fall', 'out_time_short_fall', 'duration']
                            for col in time_columns:
                                if col in df.columns:
                                    df[col] = df[col].apply(time_to_excel_format)
                            
                            cols = ['date', 'status', 'in_time_val', 'out_time_val', 
                                    'in_time_short_fall', 'out_time_short_fall', 'duration']
                            existing_cols = [col for col in cols if col in df.columns]
                            df = df[existing_cols]
                            
                            df['date_dt'] = pd.to_datetime(df['date'], format='%B %d, %Y', errors='coerce')
                            
                            # Only include rows from the correct month
                            df = df[(df['date_dt'] >= report_month_start) & (df['date_dt'] <= report_month_end)]
                            
                            all_data_frames.append(df)
                            any_data_found = True
                            logger.debug(f"Extracted {len(data_rows)} rows for {identifier} in {file_path}")
                        else:
                            log_results.append(f"[⚠️] Missing In/Out Time columns in {file_path}.")
                    else:
                        log_results.append(f"[!] Found {search_by} but no data rows in {file_path}")
                else:
                    log_results.append(f"[!] No entry found for '{identifier}' in {file_path}")
            except Exception as e:
                logger.error(f"Error processing {file_path} for {identifier}: {e}")
                log_results.append(f"[❌] Error processing {os.path.basename(file_path)}: {str(e)}")
        
        # Generate complete date range for the report (only for the detected month)
        date_range = [report_month_start + timedelta(days=x) for x in range((report_month_end - report_month_start).days + 1)]
        date_range_df = pd.DataFrame({
            'date': [d.strftime('%B %d, %Y') for d in date_range],
            'date_dt': date_range
        })
        
        if all_data_frames:
            try:
                # Combine all data frames for this employee
                combined_df = pd.concat(all_data_frames, ignore_index=True)
                combined_df['date_dt'] = pd.to_datetime(combined_df['date'], format='%B %d, %Y', errors='coerce')
                combined_df = combined_df.sort_values('date_dt').drop_duplicates(subset=['date']).dropna(subset=['date_dt'])
                
                # Merge with complete date range
                merged_df = date_range_df.merge(combined_df, on=['date', 'date_dt'], how='left')
                
                # CRITICAL FIX: Ensure status values are preserved exactly as they are
                # Create a new status column and preserve P, A, H, L values
                if 'status' not in merged_df.columns:
                    merged_df['status'] = 'A'  # Default to absent
                else:
                    # Fill NaN values with 'A'
                    merged_df['status'] = merged_df['status'].fillna('A')
                    
                    # Make sure status values are one of P, A, H, L
                    for idx, status in enumerate(merged_df['status']):
                        if status not in ['P', 'A', 'H', 'L']:
                            merged_df.at[idx, 'status'] = 'A'  # Default to absent if unrecognized
                
                # Fill missing values for time columns
                for col in ['in_time_val', 'out_time_val', 'in_time_short_fall', 'out_time_short_fall', 'duration']:
                    if col in merged_df.columns:
                        merged_df[col] = merged_df[col].fillna('')
                    else:
                        merged_df[col] = [''] * len(merged_df)
                
                # Use the merged dataframe with complete date range
                combined_df = merged_df
                
            except Exception as e:
                logger.error(f"Error combining data frames for {identifier}: {e}")
                log_results.append(f"[❌] Error combining data for {identifier}: {str(e)}")
                # Create an empty dataframe with full date range if combining fails
                combined_df = date_range_df.copy()
                combined_df['status'] = ['A'] * len(combined_df)
                combined_df['in_time_val'] = [''] * len(combined_df)
                combined_df['out_time_val'] = [''] * len(combined_df)
                combined_df['in_time_short_fall'] = [''] * len(combined_df)
                combined_df['out_time_short_fall'] = [''] * len(combined_df)
                combined_df['duration'] = [''] * len(combined_df)
        else:
            # Create an empty dataframe with full date range if no data found
            combined_df = date_range_df.copy()
            combined_df['status'] = ['A'] * len(combined_df)
            combined_df['in_time_val'] = [''] * len(combined_df)
            combined_df['out_time_val'] = [''] * len(combined_df)
            combined_df['in_time_short_fall'] = [''] * len(combined_df)
            combined_df['out_time_short_fall'] = [''] * len(combined_df)
            combined_df['duration'] = [''] * len(combined_df)
            any_data_found = True  # Mark as having data so we generate reports
        
        # Final dataframe preparation
        combined_df = combined_df.drop(columns=['date_dt'], errors='ignore')
        
        # Rename columns for the reports
        column_mapping = {
            'date': 'Date',
            'status': 'Status',
            'in_time_val': 'In Time',
            'out_time_val': 'Out Time',
            'in_time_short_fall': 'In Time Short Fall',
            'out_time_short_fall': 'Out Time Short Fall',
            'duration': 'Duration'
        }
        combined_df = combined_df.rename(columns=column_mapping)
        
        # Set up display name and sheet name
        if employee_name:
            sheet_name = employee_name[:31]
        elif employee_id:
            sheet_name = f"ID_{employee_id[:31]}"
        display_name = f"{employee_name} Att-ID:{employee_id} Designation:{designation}"
        
        # Generate reports in requested format(s)
        if output_format in ['xlsx', 'all']:
            ws = wb.create_sheet(title=sheet_name)
            
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(combined_df.columns))
            org_cell = ws.cell(row=1, column=1)
            org_cell.value = "DECCAN COLLEGE OF MEDICAL SCIENCES"
            org_cell.font = Font(bold=True, size=16)
            org_cell.alignment = Alignment(horizontal="center", vertical="center")
            org_cell.fill = PatternFill("solid", fgColor="BDD7EE")
            
            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(combined_df.columns))
            title_cell = ws.cell(row=2, column=1)
            title_cell.value = f"{department} AEBAS Attendance from {report_month_start.strftime('%B %d, %Y')} to {report_month_end.strftime('%B %d, %Y')}"
            title_cell.font = Font(bold=True, size=14)
            title_cell.alignment = Alignment(horizontal="center", vertical="center")
            title_cell.fill = PatternFill("solid", fgColor="BDD7EE")
            
            ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=len(combined_df.columns))
            info_cell = ws.cell(row=3, column=1)
            info_cell.value = display_name
            info_cell.font = Font(bold=True, size=12)
            info_cell.alignment = Alignment(horizontal="center", vertical="center")
            info_cell.fill = PatternFill("solid", fgColor="E2EFDA")
            
            for col_idx, header in enumerate(combined_df.columns, start=1):
                cell = ws.cell(row=4, column=col_idx)
                cell.value = header
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = PatternFill("solid", fgColor="F4B084")
            
            # Status colors mapping for Excel
            status_colors = {
                'P': 'C6EFCE',  # Green for Present
                'A': 'FFC7CE',  # Red for Absent
                'H': 'FFEB9C',  # Yellow for Half-day
                'L': 'DDEBF7'   # Light blue for Leave
            }
            
            for r_idx, row in enumerate(dataframe_to_rows(combined_df, index=False, header=False), start=5):
                for c_idx, value in enumerate(row, start=1):
                    cell = ws.cell(row=r_idx, column=c_idx)
                    cell.value = value
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    
                    # Apply color to status column
                    if c_idx == 2 and value in status_colors:  # Status column
                        cell.fill = PatternFill("solid", fgColor=status_colors[value])
                    
                    if c_idx in [3, 4, 5, 6, 7]:  # Time columns
                        cell.number_format = '[hh]:mm:ss'
            
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            
            for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.border = thin_border
            
            for col_idx in range(1, ws.max_column + 1):
                max_length = 0
                col_letter = get_column_letter(col_idx)
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                    for cell in row:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = max_length + 4
        
        if output_format in ['csv', 'all']:
            csv_filename = f"{sheet_name}_report.csv"
            csv_path = os.path.join(app.config['OUTPUT_FOLDER'], csv_filename)
            combined_df.to_csv(csv_path, index=False)
            output_files['csv'].append({'filename': csv_filename, 'display': display_name})
        
        if output_format in ['html', 'all']:
            html_filename = f"{sheet_name}_report.html"
            html_path = os.path.join(app.config['OUTPUT_FOLDER'], html_filename)
            
            # Formatter for status column to apply color coding
            def status_formatter(status):
                status_colors = {
                    'P': '#6BB635',  # Green for Present
                    'A': '#FF7575',  # Red for Absent
                    'H': '#FFCC66',  # Orange for Half-day
                    'L': '#A2CAED'   # Blue for Leave
                }
                
                color = status_colors.get(status, '#000000')  # Default black if status not recognized
                return f'<span style="font-weight: bold; color: {color}">{status}</span>'
            
            html_table = combined_df.to_html(
                index=False,
                classes='data',
                border=0,
                escape=False,
                formatters={'Status': status_formatter}
            )
            
            html_content = f"""
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Attendance Report - {display_name}</title>
    <style>
        @page {{
            size: A4;
            margin: 20mm;
        }}
        body {{
            font-family: 'Calibri', Arial, sans-serif;
            margin: 0;
            padding: 0;
            color: #000000;
            line-height: 1.4;
            width: 210mm;
            height: 297mm;
            box-sizing: border-box;
        }}
        .container {{
            width: 100%;
            height: 100%;
            padding: 0;
            border: 1px solid #cccccc;
            box-sizing: border-box;
        }}
        .header {{
            padding: 10px 0;
            border-bottom: 2px solid #000000;
            margin-bottom: 10px;
        }}
        .header p {{
            margin: 5px 0;
            text-align: center;
        }}
        .header .org-line {{
            font-size: 14pt;
        }}
        .header .org-name {{
            font-weight: bold;
        }}
        .header .division-line {{
            font-size: 14pt;
        }}
        .header .division-name {{
            font-weight: bold;
        }}
        .header .date-line {{
            font-size: 14pt;
        }}
        .header .employee-line {{
            font-size: 13pt;
            font-weight: bold;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            table-layout: fixed;
        }}
        th {{
            font-size: 12pt;
            padding: 4px;
            text-align: center;
            border: 1px solid #cccccc;
            background-color: #f2f2f2;
            font-weight: bold;
            text-transform: uppercase;
        }}
        td {{
            font-size: 11pt;
            padding: 4px;
            text-align: center;
            border: 1px solid #cccccc;
            word-wrap: break-word;
        }}
        tr:hover {{
            background-color: #f9f9f9;
        }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <p class="org-line">Organization: <span class="org-name">Deccan College Of Medical Sciences, Hyderabad</span></p>
            <p class="division-line">Division/Units: <span class="division-name">{department}</span></p>
            <p class="date-line">AEBAS Attendance - FROM {report_month_start.strftime('%d.%m.%Y')} TO {report_month_end.strftime('%d.%m.%Y')}</p>
            <p class="employee-line">{employee_name}   Att-ID: {employee_id}   Designation: {designation}</p>
        </div>
        {html_table}
    </div>
</body>
</html>
"""
            with open(html_path, 'w', encoding='utf-8') as f:
                f.write(html_content)
            output_files['html'].append({'filename': html_filename, 'display': display_name})
        
        display_names[identifier] = display_name
        log_results.append(f"[✅] Logs added for {display_name}")
    
    # Save Excel file if data was found and output format includes xlsx
    if any_data_found and output_format in ['xlsx', 'all']:
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            xlsx_filename = f"Employee_Reports_{timestamp}.xlsx"
            xlsx_path = os.path.join(app.config['OUTPUT_FOLDER'], xlsx_filename)
            wb.save(xlsx_path)
            output_files['xlsx'] = {'filename': xlsx_filename, 'display': 'All Employees'}
            log_results.append(f"✅ Excel report saved: {xlsx_filename}")
        except Exception as e:
            logger.error(f"Error saving Excel file: {e}")
            log_results.append(f"[❌] Error saving Excel file: {str(e)}")
    elif not any_data_found:
        log_results.append("❌ No logs found for any selected employees.")
    
    logger.info(f"Completed processing for {len(identifiers)} identifiers in {output_format} format")
    return log_results, output_files, display_names, min_date, max_date
@app.route('/')
def index():
    for file in os.listdir(app.config['UPLOAD_FOLDER']):
        if file != 'output':
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], file)
            try:
                if os.path.isfile(file_path):
                    os.unlink(file_path)
            except Exception as e:
                logger.error(f"Error deleting {file_path}: {e}")
    return render_template('index.html', departments=DEPARTMENTS)

@app.route('/upload_files', methods=['POST'])
def upload_files():
    if 'csv_files' not in request.files:
        return jsonify({"success": False, "message": "No files selected"})
    
    files = request.files.getlist('csv_files')
    department = request.form.get('department')
    if not files or files[0].filename == '':
        return jsonify({"success": False, "message": "No files selected"})
    if not department:
        return jsonify({"success": False, "message": "Please select a department"})
    
    file_paths = []
    for file in files:
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(file_path)
            file_paths.append(file_path)
    
    try:
        employees = extract_employees_from_csv(file_paths)
        return jsonify({
            "success": True,
            "message": f"Found {len(employees)} employees",
            "employee_data": employees,
            "department": department
        })
    except Exception as e:
        logger.error(f"Error processing uploaded files: {e}")
        return jsonify({
            "success": False,
            "message": f"Error processing files: {str(e)}"
        })

@app.route('/process', methods=['POST'])
def process():
    search_by = request.form.get('search_by', 'name')
    identifiers = request.form.getlist('identifiers')
    manual_identifier = request.form.get('manual_identifier', '').strip()
    department = request.form.get('department')
    
    if manual_identifier:
        identifiers.append(manual_identifier)
    if not identifiers:
        logger.warning("No identifiers provided for processing")
        return jsonify({"success": False, "message": "At least one employee must be selected"})
    
    file_paths = [os.path.join(app.config['UPLOAD_FOLDER'], f) 
                 for f in os.listdir(app.config['UPLOAD_FOLDER']) 
                 if f.endswith('.csv')]
    
    if not file_paths:
        logger.warning("No CSV files found in upload folder")
        return jsonify({"success": False, "message": "No valid CSV files found"})
    
    output_format = request.form.get('output_format', 'xlsx')
    try:
        logs, output_files, display_names, min_date, max_date = extract_employee_logs(file_paths, identifiers, search_by, output_format, department)
        return jsonify({
            "success": True,
            "logs": logs,
            "output_files": output_files,
            "display_names": display_names,
            "search_by": search_by,
            "output_format": output_format,
            "department": department,
            "min_date": min_date.strftime('%B %d, %Y') if min_date else '',
            "max_date": max_date.strftime('%B %d, %Y') if max_date else ''
        })
    except Exception as e:
        logger.error(f"Error in process endpoint: {e}")
        return jsonify({"success": False, "message": f"Error generating reports: {str(e)}"})

@app.route('/results')
def results():
    logs = request.args.get('logs', '').split('|')
    output_files = {
        'xlsx': request.args.get('xlsx') if request.args.get('xlsx') else None,
        'csv': request.args.getlist('csv[]') if request.args.getlist('csv[]') else [],
        'html': request.args.getlist('html[]') if request.args.getlist('html[]') else []
    }
    if output_files['xlsx'] and isinstance(output_files['xlsx'], str):
        output_files['xlsx'] = {'filename': output_files['xlsx'], 'display': request.args.get('xlsx_display', 'All Employees')}
    if output_files['csv'] and all(isinstance(f, str) for f in output_files['csv']):
        csv_displays = request.args.getlist('csv_display[]') if request.args.getlist('csv_display[]') else [f.split('_report')[0] for f in output_files['csv']]
        output_files['csv'] = [{'filename': f, 'display': d} for f, d in zip(output_files['csv'], csv_displays)]
    if output_files['html'] and all(isinstance(f, str) for f in output_files['html']):
        html_displays = request.args.getlist('html_display[]') if request.args.getlist('html_display[]') else [f.split('_report')[0] for f in output_files['html']]
        output_files['html'] = [{'filename': f, 'display': d} for f, d in zip(output_files['html'], html_displays)]
    display_names = {}
    search_by = request.args.get('search_by', 'name')
    output_format = request.args.get('output_format', 'xlsx')
    department = request.args.get('department', '')
    min_date = request.args.get('min_date', '')
    max_date = request.args.get('max_date', '')
    return render_template('results.html', logs=logs, output_files=output_files, display_names=display_names, search_by=search_by, output_format=output_format, department=department, min_date=min_date, max_date=max_date)

@app.route('/download/<filename>')
def download_file(filename):
    try:
        return send_file(os.path.join(app.config['OUTPUT_FOLDER'], filename), as_attachment=True)
    except Exception as e:
        logger.error(f"Error downloading file {filename}: {e}")
        flash(f"Error downloading file: {str(e)}", 'error')
        return redirect(url_for('index'))

if __name__ == '__main__':
    templates_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'templates')
    os.makedirs(templates_dir, exist_ok=True)
    
    index_html = """<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <title>Employee Log Extractor</title>
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <link href="https://fonts.googleapis.com/css2?family=Poppins:wght@300;400;500;600;700&display=swap" rel="stylesheet">
    <style>
        * {
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }
        body {
            font-family: 'Poppins', sans-serif;
            background: linear-gradient(135deg, #e0eafc, #cfdef3);
            min-height: 100vh;
            color: #2c3e50;
            overflow-x: hidden;
        }
        .container {
            max-width: 1100px;
            margin: 60px auto;
            background: #ffffff;
            padding: 45px;
            border-radius: 25px;
            box-shadow: 0 20px 50px rgba(0, 0, 0, 0.15);
            animation: fadeInUp 0.6s ease-out;
        }
        @keyframes fadeInUp {
            from { opacity: 0; transform: translateY(30px); }
            to { opacity: 1; transform: translateY(0); }
        }
        h1 {
            text-align: center;
            color: #34495e;
            font-weight: 700;
            font-size: 3rem;
            margin-bottom: 40px;
            text-transform: uppercase;
            letter-spacing: 2px;
            animation: bounceIn 0.8s ease;
        }
        @keyframes bounceIn {
            0% { transform: scale(0.9); opacity: 0; }
            50% { transform: scale(1.05); opacity: 0.5; }
            100% { transform: scale(1); opacity: 1; }
        }
        .form-group {
            margin-bottom: 35px;
        }
        label {
            font-weight: 600;
            margin-bottom: 12px;
            display: block;
            font-size: 1.2rem;
            color: #34495e;
            transition: color 0.3s ease;
        }
        label:hover {
            color: #2980b9;
        }
        input[type="text"],
        select {
            width: 100%;
            padding: 16px 20px;
            font-size: 1.1rem;
            border: 2px solid #ecf0f1;
            border-radius: 15px;
            background: #f9fafb;
            transition: all 0.4s ease;
            box-shadow: inset 0 2px 5px rgba(0, 0, 0, 0.05);
        }
        input[type="text"]:focus,
        select:focus {
            border-color: #3498db;
            background: #fff;
            outline: none;
            box-shadow: 0 0 15px rgba(52, 152, 219, 0.4), inset 0 2px 5px rgba(0, 0, 0, 0.05);
        }
        .search-options {
            display: flex;
            gap: 25px;
            flex-wrap: wrap;
            margin-top: 12px;
        }
        .search-option {
            display: flex;
            align-items: center;
            gap: 12px;
            font-size: 1rem;
        }
        input[type="radio"] {
            display: none;
        }
        .search-option label {
            cursor: pointer;
            padding-left: 32px;
            position: relative;
            user-select: none;
            font-weight: 500;
            transition: color 0.3s ease;
        }
        .search-option label:hover {
            color: #2980b9;
        }
        .search-option label::before {
            content: '';
            position: absolute;
            left: 0;
            top: 50%;
            transform: translateY(-50%);
            width: 22px;
            height: 22px;
            border: 2px solid #95a5a6;
            border-radius: 50%;
            background: #fff;
            transition: border-color 0.3s ease;
        }
        .search-option label::after {
            content: '';
            position: absolute;
            left: 7px;
            top: 50%;
            transform: translateY(-50%) scale(0);
            width: 10px;
            height: 10px;
            border-radius: 50%;
            background: #3498db;
            transition: transform 0.3s ease;
        }
        input[type="radio"]:checked + label::before {
            border-color: #3498db;
        }
        input[type="radio"]:checked + label::after {
            transform: translateY(-50%) scale(1);
        }
        .file-input {
            border: 2px dashed #bdc3c7;
            padding: 50px;
            text-align: center;
            background-color: #f9fafb;
            border-radius: 18px;
            cursor: pointer;
            transition: all 0.4s ease;
            animation: pulse 2s infinite;
        }
        @keyframes pulse {
            0% { transform: scale(1); }
            50% { transform: scale(1.02); }
            100% { transform: scale(1); }
        }
        .file-input:hover {
            border-color: #3498db;
            background-color: #eef6ff;
            box-shadow: 0 10px 25px rgba(52, 152, 219, 0.3);
        }
        input[type="file"] {
            display: none;
        }
        .selected-files {
            margin-top: 15px;
            font-size: 1rem;
            color: #7f8c8d;
            text-align: center;
            background: #f1f3f5;
            padding: 12px;
            border-radius: 12px;
            box-shadow: 0 2px 10px rgba(0, 0, 0, 0.05);
            animation: fadeIn 0.5s ease;
        }
        button {
            background: linear-gradient(90deg, #3498db, #2980b9);
            color: white;
            padding: 16px 32px;
            border: none;
            border-radius: 15px;
            font-size: 1.1rem;
            font-weight: 600;
            cursor: pointer;
            box-shadow: 0 5px 15px rgba(52, 152, 219, 0.3);
            transition: all 0.4s ease;
        }
        button:hover {
            background: linear-gradient(90deg, #2980b9, #2471a3);
            transform: translateY(-3px);
            box-shadow: 0 8px 20px rgba(52, 152, 219, 0.4);
        }
        button:disabled {
            background: #95a5a6;
            cursor: not-allowed;
            opacity: 0.7;
        }
        .flash-messages {
            margin-bottom: 30px;
        }
        .flash-message {
            padding: 15px;
            border-radius: 12px;
            font-size: 1rem;
            margin-bottom: 10px;
            animation: slideInLeft 0.5s ease;
        }
        @keyframes slideInLeft {
            from { transform: translateX(-20px); opacity: 0; }
            to { transform: translateX(0); opacity: 1; }
        }
        .error {
            background-color: #fce4e4;
            color: #c0392b;
            border: 1px solid #e74c3c;
        }
        #upload-section, #employee-selection-section, #processing-section {
            transition: all 0.5s ease;
        }
        #employee-selection-section, #processing-section {
            display: none;
        }
        .employee-list {
            max-height: 450px;
            overflow-y: auto;
            border-radius: 15px;
            background: linear-gradient(135deg, #ffffff, #f9fafb);
            box-shadow: 0 10px 25px rgba(0, 0, 0, 0.08);
            border: 1px solid #ecf0f1;
            animation: fadeIn 0.5s ease;
        }
        .employee-item {
            padding: 18px 28px;
            border-bottom: 1px solid rgba(236, 240, 241, 0.5);
            cursor: pointer;
            display: flex;
            justify-content: space-between;
            align-items: center;
            transition: all 0.4s ease;
            background: #fff;
            animation: slideUp 0.3s ease;
        }
        @keyframes slideUp {
            from { transform: translateY(10px); opacity: 0; }
            to { transform: translateY(0); opacity: 1; }
        }
        .employee-item:hover {
            background: linear-gradient(90deg, #eef6ff, #e9f2ff);
            transform: translateX(8px);
            box-shadow: 0 2px 10px rgba(52, 152, 219, 0.1);
        }
        .employee-item.selected {
            background: #d4e6ff;
            font-weight: 500;
        }
        .employee-name {
            font-weight: 600;
            color: #34495e;
            transition: color 0.3s ease;
        }
        .employee-id {
            font-size: 0.9rem;
            color: #95a5a6;
            padding: 6px 12px;
            border-radius: 12px;
            background: #f1f3f5;
            transition: background 0.3s ease;
        }
        .employee-item:hover .employee-name {
            color: #2980b9;
        }
        .employee-item:hover .employee-id {
            background: #d4e6ff;
        }
        .employee-search {
            margin-bottom: 18px;
            position: relative;
        }
        .section-header {
            display: flex;
            justify-content: space-between;
            align-items: center;
            margin-bottom: 20px;
            border-bottom: 2px solid #ecf0f1;
            padding-bottom: 12px;
            animation: fadeIn 0.5s ease;
        }
        .info-text {
            background: linear-gradient(135deg, #eef6ff, #e9f2ff);
            border: 1px solid #d4e6ff;
            border-radius: 15px;
            padding: 18px;
            margin-bottom: 30px;
            font-size: 1rem;
            color: #2980b9;
            box-shadow: 0 4px 15px rgba(52, 152, 219, 0.15);
            animation: fadeIn 0.5s ease;
        }
        .loading {
            text-align: center;
            padding: 30px;
            color: #7f8c8d;
            display: flex;
            align-items: center;
            justify-content: center;
            animation: pulse 2s infinite;
        }
        .spinner {
            border: 6px solid #ecf0f1;
            width: 40px;
            height: 40px;
            border-radius: 50%;
            border-left-color: #3498db;
            animation: spin 1.2s linear infinite;
            margin-right: 20px;
        }
        @keyframes spin {
            0% { transform: rotate(0deg); }
            100% { transform: rotate(360deg); }
        }
        .back-button {
            background: linear-gradient(90deg, #95a5a6, #7f8c8d);
            margin-right: 15px;
        }
        .button-row {
            display: flex;
            gap: 20px;
            flex-wrap: wrap;
            justify-content: center;
            animation: fadeIn 0.5s ease;
        }
        .selected-employees {
            margin-bottom: 25px;
            padding: 18px;
            background: #eef6ff;
            border-radius: 15px;
            box-shadow: 0 4px 15px rgba(52, 152, 219, 0.15);
            animation: slideInRight 0.5s ease;
        }
        @keyframes slideInRight {
            from { transform: translateX(20px); opacity: 0; }
            to { transform: translateX(0); opacity: 1; }
        }
        .selected-employees ul {
            list-style: none;
            padding: 0;
        }
        .selected-employees li {
            padding: 8px 0;
            color: #34495e;
            transition: color 0.3s ease;
            animation: fadeIn 0.5s ease backwards;
        }
        .selected-employees li:hover {
            color: #2980b9;
        }
        footer {
            text-align: center;
            margin-top: 50px;
            padding: 25px 0;
            color: #7f8c8d;
            font-size: 1rem;
            border-top: 1px solid #ecf0f1;
            animation: fadeIn 0.5s ease;
        }
        .manual-search {
            margin-bottom: 18px;
            position: relative;
        }
        .add-button {
            position: absolute;
            right: 10px;
            top: 50%;
            transform: translateY(-50%);
            background: linear-gradient(90deg, #2ecc71, #27ae60);
            color: white;
            padding: 8px 16px;
            border: none;
            border-radius: 12px;
            font-size: 0.9rem;
            font-weight: 600;
            cursor: pointer;
            box-shadow: 0 3px 10px rgba(46, 204, 113, 0.3);
            transition: all 0.4s ease;
        }
        .add-button:hover {
            background: linear-gradient(90deg, #27ae60, #219653);
            transform: translateY(-53%);
            box-shadow: 0 5px 15px rgba(46, 204, 113, 0.4);
        }
        .add-button:disabled {
            background: #95a5a6;
            cursor: not-allowed;
            opacity: 0.7;
        }
        .select-all-btn {
            background: linear-gradient(90deg, #2ecc71, #27ae60);
            margin-bottom: 18px;
            display: inline-block;
            animation: bounceIn 0.8s ease;
        }
        .select-all-btn:hover {
            background: linear-gradient(90deg, #27ae60, #219653);
            transform: translateY(-3px);
        }
        .download-options {
            margin-top: 18px;
        }
        select {
            appearance: none;
            background: url('data:image/svg+xml;utf8,<svg xmlns="http://www.w3.org/2000/svg" width="24" height="24" fill="%233498db"><path d="M7 10l5 5 5-5z"/></svg>') no-repeat right 15px center;
            padding-right: 40px;
            cursor: pointer;
        }
        select:focus {
            border-color: #2980b9;
        }
    </style>
</head>
<body>
    <div class="container">
        <h1>Employee Log Extractor</h1>
        {% with messages = get_flashed_messages(with_categories=true) %}
            {% if messages %}
                <div class="flash-messages">
                    {% for category, message in messages %}
                        <div class="flash-message {{ category }}">{{ message }}</div>
                    {% endfor %}
                </div>
            {% endif %}
        {% endwith %}
        <div id="upload-section">
            <div class="info-text">
                <strong>Step 1:</strong> Upload your CSV files and select a department to begin extracting employee logs.
            </div>
            <div class="form-group">
                <label for="department">Select Department</label>
                <select id="department" name="department">
                    <option value="">Select a department</option>
                    {% for dept in departments %}
                        <option value="{{ dept }}">{{ dept }}</option>
                    {% endfor %}
                </select>
            </div>
            <div class="form-group">
                <label for="csv_files">Upload CSV Files</label>
                <div class="file-input" id="file-drop-area">
                    <p>Drag & drop CSV files here or click to browse</p>
                    <input type="file" id="csv_files" name="csv_files" accept=".csv" multiple>
                </div>
                <div class="selected-files" id="selected-files">
                    No files selected
                </div>
            </div>
            <button id="upload-button" type="button">Upload Files</button>
        </div>
        <div id="employee-selection-section">
            <div class="info-text">
                <strong>Step 2:</strong> Select or search employees to generate reports.
            </div>
            <div class="form-group">
                <div class="section-header">
                    <label>Employee List</label>
                    <span id="employee-count">0 employees found</span>
                </div>
                <div class="employee-search">
                    <input type="text" id="employee-search-input" placeholder="Search by name or ID...">
                </div>
                <button class="select-all-btn" id="select-all-button" type="button">Select All</button>
                <div class="employee-list" id="employee-list"></div>
            </div>
            <div class="manual-search form-group">
                <label for="manual_identifier">Manual Search (Name or ID)</label>
                <input type="text" id="manual_identifier" name="manual_identifier" placeholder="Enter name or ID...">
                <button class="add-button" id="add-manual-button" type="button" disabled>Add</button>
            </div>
            <div class="selected-employees" id="selected-employees" style="display: none;">
                <strong>Selected Employees:</strong>
                <ul id="selected-employees-list"></ul>
            </div>
            <div class="form-group">
                <label for="search_by">Search By</label>
                <div class="search-options">
                    <div class="search-option">
                        <input type="radio" id="search_by_name" name="search_by" value="name" checked>
                        <label for="search_by_name">Employee Name</label>
                    </div>
                    <div class="search-option">
                        <input type="radio" id="search_by_id" name="search_by" value="id">
                        <label for="search_by_id">Attendance ID</label>
                    </div>
                </div>
            </div>
            <div class="form-group download-options">
                <label for="output_format">Download Format</label>
                <select id="output_format" name="output_format">
                    <option value="xlsx">Excel</option>
                    <option value="csv">CSV</option>
                    <option value="html">HTML</option>
                    <option value="all">All Formats</option>
                </select>
            </div>
            <div class="button-row">
                <button class="back-button" id="back-to-upload-button" type="button">Back to Upload</button>
                <button id="generate-reports-button" type="button" disabled>Generate Reports</button>
            </div>
        </div>
        <div id="processing-section" style="display: none;">
            <div class="info-text">
                <strong>Processing...</strong> Please wait while we prepare your reports.
            </div>
            <div class="loading">
                <div class="spinner"></div>
                <span>Processing files...</span>
            </div>
        </div>
        <footer>
            © 2025 Employee Log Extractor | Developed by Mir Abdul Aziz Khan
        </footer>
    </div>
    <script>
        document.addEventListener('DOMContentLoaded', function() {
            const uploadSection = document.getElementById('upload-section');
            const employeeSelectionSection = document.getElementById('employee-selection-section');
            const processingSection = document.getElementById('processing-section');
            const fileInput = document.getElementById('csv_files');
            const fileDropArea = document.getElementById('file-drop-area');
            const selectedFilesDisplay = document.getElementById('selected-files');
            const uploadButton = document.getElementById('upload-button');
            const employeeList = document.getElementById('employee-list');
            const employeeCount = document.getElementById('employee-count');
            const employeeSearchInput = document.getElementById('employee-search-input');
            const backToUploadButton = document.getElementById('back-to-upload-button');
            const generateReportsButton = document.getElementById('generate-reports-button');
            const selectedEmployeesDisplay = document.getElementById('selected-employees');
            const selectedEmployeesList = document.getElementById('selected-employees-list');
            const searchByName = document.getElementById('search_by_name');
            const searchById = document.getElementById('search_by_id');
            const manualIdentifierInput = document.getElementById('manual_identifier');
            const addManualButton = document.getElementById('add-manual-button');
            const selectAllButton = document.getElementById('select-all-button');
            const outputFormatSelect = document.getElementById('output_format');
            const departmentSelect = document.getElementById('department');
            
            let employeesData = [];
            let selectedEmployees = [];
            let manualEmployees = [];
            let selectedDepartment = '';
            
            fileInput.addEventListener('change', handleFileSelect);
            fileDropArea.addEventListener('dragover', function(e) {
                e.preventDefault();
                fileDropArea.style.borderColor = '#3498db';
                fileDropArea.style.backgroundColor = '#eef6ff';
            });
            fileDropArea.addEventListener('dragleave', function(e) {
                e.preventDefault();
                fileDropArea.style.borderColor = '#bdc3c7';
                fileDropArea.style.backgroundColor = '#f9fafb';
            });
            fileDropArea.addEventListener('drop', function(e) {
                e.preventDefault();
                fileDropArea.style.borderColor = '#bdc3c7';
                fileDropArea.style.backgroundColor = '#f9fafb';
                fileInput.files = e.dataTransfer.files;
                handleFileSelect();
            });
            fileDropArea.addEventListener('click', function() {
                fileInput.click();
            });
            uploadButton.addEventListener('click', handleUpload);
            backToUploadButton.addEventListener('click', function() {
                employeeSelectionSection.style.display = 'none';
                uploadSection.style.display = 'block';
            });
            generateReportsButton.addEventListener('click', handleGenerateReports);
            employeeSearchInput.addEventListener('input', filterEmployees);
            selectAllButton.addEventListener('click', toggleSelectAll);
            manualIdentifierInput.addEventListener('input', updateManualState);
            addManualButton.addEventListener('click', addManualEmployee);
            
            function handleFileSelect() {
                const files = fileInput.files;
                if (files.length > 0) {
                    let fileNames = [];
                    let validFiles = true;
                    for (let i = 0; i < files.length; i++) {
                        if (!files[i].name.toLowerCase().endsWith('.csv')) {
                            validFiles = false;
                        }
                        fileNames.push(files[i].name);
                    }
                    if (!validFiles) {
                        selectedFilesDisplay.innerHTML = '<span style="color: #c0392b;">Please select only CSV files</span>';
                        return;
                    }
                    selectedFilesDisplay.innerHTML = `Selected ${files.length} file${files.length > 1 ? 's' : ''}: ${fileNames.join(', ')}`;
                } else {
                    selectedFilesDisplay.innerHTML = 'No files selected';
                }
            }
            
            function handleUpload() {
                const files = fileInput.files;
                if (files.length === 0) {
                    alert('Please select at least one CSV file');
                    return;
                }
                if (!departmentSelect.value) {
                    alert('Please select a department');
                    return;
                }
                const formData = new FormData();
                for (let i = 0; i < files.length; i++) {
                    formData.append('csv_files', files[i]);
                }
                formData.append('department', departmentSelect.value);
                uploadButton.disabled = true;
                uploadButton.textContent = 'Uploading...';
                fetch('/upload_files', {
                    method: 'POST',
                    body: formData
                })
                .then(response => response.json())
                .then(data => {
                    uploadButton.disabled = false;
                    uploadButton.textContent = 'Upload Files';
                    if (data.success) {
                        uploadSection.style.display = 'none';
                        employeeSelectionSection.style.display = 'block';
                        employeeCount.textContent = `${data.employee_data.length} employee${data.employee_data.length !== 1 ? 's' : ''} found`;
                        employeesData = data.employee_data;
                        selectedDepartment = data.department;
                        displayEmployees(employeesData);
                    } else {
                        alert(data.message);
                    }
                })
                .catch(error => {
                    uploadButton.disabled = false;
                    uploadButton.textContent = 'Upload Files';
                    alert('Error uploading files: ' + error.message);
                });
            }
            
            function displayEmployees(employees) {
                employeeList.innerHTML = '';
                if (!employees || employees.length === 0) {
                    employeeList.innerHTML = '<div class="employee-item"><span class="employee-name">No employees found</span></div>';
                    return;
                }
                employees.forEach(employee => {
                    const employeeItem = document.createElement('div');
                    employeeItem.className = 'employee-item';
                    if (selectedEmployees.some(e => e.id === employee.id && e.name === employee.name)) {
                        employeeItem.classList.add('selected');
                    }
                    const nameSpan = document.createElement('span');
                    nameSpan.className = 'employee-name';
                    nameSpan.textContent = employee.name || 'Unnamed Employee';
                    const idSpan = document.createElement('span');
                    idSpan.className = 'employee-id';
                    idSpan.textContent = employee.id ? `(ID: ${employee.id})` : '(No ID)';
                    employeeItem.appendChild(nameSpan);
                    employeeItem.appendChild(idSpan);
                    employeeItem.dataset.name = employee.name;
                    employeeItem.dataset.id = employee.id;
                    employeeItem.dataset.display = employee.display;
                    employeeItem.addEventListener('click', function() {
                        toggleEmployeeSelection(employee, employeeItem);
                    });
                    employeeList.appendChild(employeeItem);
                });
            }
            
            function toggleEmployeeSelection(employee, element) {
                const index = selectedEmployees.findIndex(e => e.id === employee.id && e.name === employee.name);
                if (index === -1) {
                    selectedEmployees.push(employee);
                    element.classList.add('selected');
                } else {
                    selectedEmployees.splice(index, 1);
                    element.classList.remove('selected');
                }
                updateSelectedEmployeesDisplay();
                updateButtonState();
                if (selectedEmployees.length > 0) {
                    manualIdentifierInput.disabled = true;
                    addManualButton.disabled = true;
                    manualIdentifierInput.value = '';
                    manualEmployees = [];
                } else {
                    manualIdentifierInput.disabled = false;
                }
            }
            
            function toggleSelectAll() {
                if (selectedEmployees.length === employeesData.length) {
                    selectedEmployees = [];
                    document.querySelectorAll('.employee-item').forEach(item => item.classList.remove('selected'));
                } else {
                    selectedEmployees = [...employeesData];
                    document.querySelectorAll('.employee-item').forEach(item => item.classList.add('selected'));
                }
                updateSelectedEmployeesDisplay();
                updateButtonState();
                if (selectedEmployees.length > 0) {
                    manualIdentifierInput.disabled = true;
                    addManualButton.disabled = true;
                    manualIdentifierInput.value = '';
                    manualEmployees = [];
                } else {
                    manualIdentifierInput.disabled = false;
                }
            }
            
            function updateSelectedEmployeesDisplay() {
                selectedEmployeesList.innerHTML = '';
                if (selectedEmployees.length > 0 || manualEmployees.length > 0) {
                    selectedEmployeesDisplay.style.display = 'block';
                    [...selectedEmployees, ...manualEmployees.map(emp => ({ display: emp }))].forEach((emp, index) => {
                        const li = document.createElement('li');
                        li.textContent = emp.display;
                        li.style.animationDelay = `${index * 0.1}s`;
                        selectedEmployeesList.appendChild(li);
                    });
                    generateReportsButton.disabled = false;
                } else {
                    selectedEmployeesDisplay.style.display = 'none';
                    generateReportsButton.disabled = true;
                }
            }
            
            function filterEmployees() {
                const searchTerm = employeeSearchInput.value.toLowerCase();
                const filteredEmployees = employeesData.filter(employee => {
                    const nameMatch = employee.name && employee.name.toLowerCase().includes(searchTerm);
                    const idMatch = employee.id && employee.id.toLowerCase().includes(searchTerm);
                    return nameMatch || idMatch;
                });
                displayEmployees(filteredEmployees);
            }
            
            function updateManualState() {
                addManualButton.disabled = !manualIdentifierInput.value.trim();
                if (manualIdentifierInput.value.trim() && selectedEmployees.length === 0) {
                    manualIdentifierInput.disabled = false;
                }
            }
            
            function addManualEmployee() {
                const identifier = manualIdentifierInput.value.trim();
                if (identifier && !manualEmployees.includes(identifier)) {
                    manualEmployees.push(identifier);
                    manualIdentifierInput.value = '';
                    addManualButton.disabled = true;
                    updateSelectedEmployeesDisplay();
                    employeeSearchInput.disabled = true;
                    selectAllButton.disabled = true;
                    selectedEmployees = [];
                    displayEmployees(employeesData);
                }
            }
            
            function updateButtonState() {
                generateReportsButton.disabled = selectedEmployees.length === 0 && manualEmployees.length === 0;
            }
            
            function handleGenerateReports() {
                if (selectedEmployees.length === 0 && manualEmployees.length === 0) {
                    alert('Please select at least one employee or enter a manual identifier');
                    return;
                }
                const formData = new FormData();
                formData.append('search_by', searchByName.checked ? 'name' : 'id');
                formData.append('output_format', outputFormatSelect.value);
                formData.append('department', selectedDepartment);
                selectedEmployees.forEach(emp => {
                    formData.append('identifiers', searchByName.checked ? emp.name : emp.id);
                });
                manualEmployees.forEach(emp => {
                    formData.append('identifiers', emp);
                });
                employeeSelectionSection.style.display = 'none';
                processingSection.style.display = 'block';
                generateReportsButton.disabled = true;
                
                const controller = new AbortController();
                const timeoutId = setTimeout(() => controller.abort(), 60000);
                
                fetch('/process', {
                    method: 'POST',
                    body: formData,
                    signal: controller.signal
                })
                .then(response => {
                    clearTimeout(timeoutId);
                    if (!response.ok) {
                        throw new Error(`HTTP error! status: ${response.status}`);
                    }
                    return response.json();
                })
                .then(data => {
                    if (data.success) {
                        const csvFiles = data.output_files.csv.map(f => `csv[]=${encodeURIComponent(f.filename)}&csv_display[]=${encodeURIComponent(f.display)}`).join('&');
                        const htmlFiles = data.output_files.html.map(f => `html[]=${encodeURIComponent(f.filename)}&html_display[]=${encodeURIComponent(f.display)}`).join('&');
                        const xlsxFile = data.output_files.xlsx ? `xlsx=${encodeURIComponent(data.output_files.xlsx.filename)}&xlsx_display=${encodeURIComponent(data.output_files.xlsx.display)}` : '';
                        const url = `/results?logs=${encodeURIComponent(data.logs.join('|'))}&${xlsxFile}&${csvFiles}&${htmlFiles}&search_by=${encodeURIComponent(data.search_by)}&output_format=${encodeURIComponent(data.output_format)}&department=${encodeURIComponent(data.department)}&min_date=${encodeURIComponent(data.min_date)}&max_date=${encodeURIComponent(data.max_date)}`;
                        window.location.href = url;
                    } else {
                        throw new Error(data.message);
                    }
                })
                .catch(error => {
                    console.error('Fetch error:', error);
                    processingSection.style.display = 'none';
                    employeeSelectionSection.style.display = 'block';
                    generateReportsButton.disabled = false;
                    alert('Error generating reports: ' + (error.message || 'Unknown error'));
                });
            }
            
            employeeList.innerHTML = '<div class="employee-item"><span class="employee-name">Upload files to see employees</span></div>';
        });
    </script>
</body>
</html>

"""
    
    with open(os.path.join(templates_dir, 'index.html'), 'w', encoding='utf-8') as f:
        f.write(index_html)
    
    results_html = """<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <title>Employee Log Processing Results</title>
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <link href="https://fonts.googleapis.com/css2?family=Poppins:wght@300;400;500;600;700&display=swap" rel="stylesheet">
    <style>
        * {
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }
        body {
            font-family: 'Poppins', sans-serif;
            background: linear-gradient(135deg, #e0eafc, #cfdef3);
            min-height: 100vh;
            color: #2c3e50;
        }
        .container {
            max-width: 1100px;
            margin: 60px auto;
            background: #ffffff;
            padding: 45px;
            border-radius: 25px;
            box-shadow: 0 20px 50px rgba(0, 0, 0, 0.15);
            animation: fadeInUp 0.6s ease-out;
        }
        h1 {
            text-align: center;
            color: #34495e;
            font-weight: 700;
            font-size: 3rem;
            margin-bottom: 40px;
            text-transform: uppercase;
            letter-spacing: 2px;
            animation: bounceIn 0.8s ease;
        }
        .header-box {
            background: linear-gradient(135deg, #e9f7e8, #d9e9d6);
            border: 1px solid #c8e6c9;
            border-radius: 18px;
            padding: 25px;
            margin-bottom: 40px;
            text-align: center;
            box-shadow: 0 4px 15px rgba(200, 230, 201, 0.3);
            animation: fadeIn 0.5s ease;
        }
        .header-box h2 {
            color: #34495e;
            margin: 0 0 12px 0;
            font-weight: 600;
            font-size: 1.8rem;
        }
        .header-box p {
            color: #7f8c8d;
            font-size: 1rem;
            margin: 0;
        }
        h3 {
            color: #34495e;
            font-weight: 600;
            font-size: 1.5rem;
            margin-bottom: 18px;
            border-bottom: 2px solid #ecf0f1;
            padding-bottom: 8px;
            animation: fadeIn 0.5s ease;
        }
        .log-container {
            background-color: #f9fafb;
            border: 1px solid #ecf0f1;
            border-radius: 15px;
            padding: 18px;
            margin-bottom: 40px;
            font-size: 0.9rem;
            max-height: 300px;
            overflow-y: auto;
            box-shadow: 0 4px 15px rgba(0, 0, 0, 0.05);
            animation: slideInUp 0.5s ease;
        }
        .log-line {
            margin-bottom: 8px;
            line-height: 1.5;
            transition: transform 0.3s ease, color 0.3s ease;
        }
        .log-line:hover {
            transform: translateX(8px);
            color: #2980b9;
        }
        .success-line {
            color: #27ae60;
        }
        .error-line {
            color: #c0392b;
        }
        .warning-line {
            color: #e67e22;
        }
        .download-section {
            margin-top: 40px;
            background: linear-gradient(135deg, #eef6ff, #e9f2ff);
            border: 1px solid #d4e6ff;
            border-radius: 18px;
            padding: 25px;
            box-shadow: 0 4px 15px rgba(52, 152, 219, 0.15);
            animation: fadeIn 0.5s ease;
        }
        .download-header {
            margin: 0 0 20px 0;
            color: #2980b9;
            font-weight: 600;
            font-size: 1.4rem;
        }
        .download-buttons {
            display: grid;
            grid-template-columns: repeat(auto-fill, minmax(240px, 1fr));
            gap: 25px;
            animation: fadeIn 0.5s ease;
        }
        .download-button {
            background: linear-gradient(90deg, #3498db, #2980b9);
            color: white;
            padding: 16px 20px;
            border-radius: 15px;
            text-decoration: none;
            font-weight: 600;
            display: flex;
            align-items: center;
            justify-content: center;
            transition: all 0.4s ease;
            box-shadow: 0 5px 15px rgba(52, 152, 219, 0.3);
            white-space: nowrap;
            overflow: hidden;
            text-overflow: ellipsis;
        }
        .download-button:hover {
            background: linear-gradient(90deg, #2980b9, #2471a3);
            transform: translateY(-3px);
            box-shadow: 0 8px 20px rgba(52, 152, 219, 0.4);
        }
        .buttons {
            margin-top: 40px;
            display: flex;
            gap: 25px;
            flex-wrap: wrap;
            justify-content: center;
            animation: fadeIn 0.5s ease;
        }
        .button {
            background: linear-gradient(90deg, #3498db, #2980b9);
            color: white;
            padding: 16px 32px;
            border-radius: 15px;
            font-size: 1.1rem;
            font-weight: 600;
            text-decoration: none;
            text-align: center;
            min-width: 180px;
            transition: all 0.4s ease;
            box-shadow: 0 5px 15px rgba(52, 152, 219, 0.3);
        }
        .button:hover {
            background: linear-gradient(90deg, #2980b9, #2471a3);
            transform: translateY(-3px);
            box-shadow: 0 8px 20px rgba(52, 152, 219, 0.4);
        }
        .button-back {
            background: linear-gradient(90deg, #95a5a6, #7f8c8d);
        }
        .button-back:hover {
            background: linear-gradient(90deg, #7f8c8d, #6c757d);
        }
        footer {
            text-align: center;
            margin-top: 50px;
            padding: 25px 0;
            color: #7f8c8d;
            font-size: 1rem;
            border-top: 1px solid #ecf0f1;
            animation: fadeIn 0.5s ease;
        }
    </style>
</head>
<body>
    <div class="container">
        <h1>Processing Results</h1>
        <div class="header-box">
            <h2>{{ department }} Employees</h2>
            <p>Search method: {% if search_by == 'name' %}By Name{% else %}By ID{% endif %} | Format: {{ output_format|upper }}</p>
        </div>
        <h3>Processing Log</h3>
        <div class="log-container">
            {% for log in logs %}
                <div class="log-line 
                    {% if '[✅]' in log or '✅' in log %}success-line{% endif %}
                    {% if '[❌]' in log or '❌' in log %}error-line{% endif %}
                    {% if '[⚠️]' in log or '[!]' in log %}warning-line{% endif %}">
                    {{ log }}
                </div>
            {% endfor %}
        </div>
        {% if output_files[output_format] or output_format == 'all' %}
            <div class="download-section">
                <h3 class="download-header">Download Reports</h3>
                {% if output_files.xlsx %}
                    <h4>Excel Report</h4>
                    <div class="download-buttons">
                        <a href="{{ url_for('download_file', filename=output_files.xlsx.filename) }}" class="download-button" title="{{ output_files.xlsx.display }}">
                            {{ output_files.xlsx.display }}
                        </a>
                    </div>
                {% endif %}
                {% if output_files.csv or (output_format == 'all' and output_files.csv) %}
                    <h4>CSV Reports</h4>
                    <div class="download-buttons">
                        {% for file in output_files.csv %}
                            <a href="{{ url_for('download_file', filename=file.filename) }}" class="download-button" title="{{ file.display }}">
                                {{ file.display }}
                            </a>
                        {% endfor %}
                    </div>
                {% endif %}
                {% if output_files.html or (output_format == 'all' and output_files.html) %}
                    <h4>HTML Reports</h4>
                    <div class="download-buttons">
                        {% for file in output_files.html %}
                            <a href="{{ url_for('download_file', filename=file.filename) }}" class="download-button" title="{{ file.display }}">
                                {{ file.display }}
                            </a>
                        {% endfor %}
                    </div>
                {% endif %}
            </div>
        {% endif %}
        <div class="buttons">
            <a href="{{ url_for('index') }}" class="button button-back">Back to Home</a>
        </div>
        <footer>
            © 2025 Employee Log Extractor | Developed by Mir Abdul Aziz Khan
        </footer>
    </div>
</body>
</html>
"""
    
    with open(os.path.join(templates_dir, 'results.html'), 'w', encoding='utf-8') as f:
        f.write(results_html)
    
    app.run(debug=True, host='0.0.0.0', port=5000)